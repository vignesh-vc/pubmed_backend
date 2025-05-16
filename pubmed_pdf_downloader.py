import os
import time
import glob
import shutil
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def process_excel(input_excel_path, output_excel_path):
    # Load Excel
    try:
        df = pd.read_excel(input_excel_path)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return

    if 'Comment' not in df.columns:
        df['Comment'] = ''

    if 'Result Count' not in df.columns:
        df['Result Count'] = ''

    # Setup Chrome Download Directory
    download_dir = os.path.join(os.getcwd(), "downloads")
    os.makedirs(download_dir, exist_ok=True)

    chrome_options = Options()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    chrome_options.add_argument("--start-maximized")

    # Launch Chrome
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 20)

    try:
        for index, row in df.iterrows():
            search_query = row.get('Search Strategy 1')
            target_filename = row.get('File Name')

            if pd.isna(search_query) or pd.isna(target_filename):
                print(f"⏭️ Skipping row {index + 2} due to empty values")
                df.at[index, 'Comment'] = 'Empty values'
                continue

            print(f"\n🔍 Searching for: {search_query}")
            driver.get("https://pubmed.ncbi.nlm.nih.gov/")

            try:
                search_box = wait.until(EC.presence_of_element_located((By.ID, "id_term")))
                search_box.clear()
                search_box.send_keys(search_query)
                search_box.send_keys(Keys.ENTER)

                wait.until(EC.presence_of_element_located((By.CLASS_NAME, "results-amount")))
                time.sleep(2)

                result_text = driver.find_element(By.CLASS_NAME, "results-amount").text
                result_count = int(result_text.split()[0].replace(",", ""))
                df.at[index, 'Result Count'] = result_count

                if result_count > 1000:
                    print(f"⚠️ Too many results ({result_count}) — skipping download")
                    df.at[index, 'Comment'] = 'Too many results'
                    continue

                # Quoted warning
                warnings = driver.find_elements(By.CLASS_NAME, "usa-alert-body")
                for w in warnings:
                    if "Quoted phrase not found in" in w.text:
                        print("⚠️ Quoted phrase warning — skipping")
                        df.at[index, 'Comment'] = 'Quoted phrase warning'
                        raise Exception("Quoted phrase issue")

            except Exception as e:
                if not df.at[index, 'Comment']:
                    df.at[index, 'Comment'] = 'Search/warning error'
                print(f"⚠️ Search error: {e}")
                continue

            try:
                save_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Save')]")))
                save_btn.click()

                wait.until(EC.visibility_of_element_located((By.ID, "save-action-selection")))
                driver.find_element(By.ID, "save-action-selection").send_keys(Keys.DOWN + Keys.RETURN)
                time.sleep(0.5)
                driver.find_element(By.ID, "save-action-format").send_keys(Keys.DOWN + Keys.RETURN)
                time.sleep(0.5)

                create_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Create file')]")))
                create_btn.click()

                print("⏳ Waiting for file to download...")
                time.sleep(7)

                list_of_files = glob.glob(os.path.join(download_dir, "pubmed-*.*"))
                if list_of_files:
                    latest_file = max(list_of_files, key=os.path.getctime)
                    new_path = os.path.join(download_dir, f"{target_filename}.txt")
                    shutil.move(latest_file, new_path)
                    print(f"✅ Downloaded: {target_filename}.txt")
                    df.at[index, 'Comment'] = 'Downloaded'
                else:
                    print("❌ No downloaded file found.")
                    df.at[index, 'Comment'] = 'Download error'

            except Exception as e:
                print(f"⚠️ Download failed: {e}")
                df.at[index, 'Comment'] = 'Download error'
                continue

    finally:
        driver.quit()
        print("\n🛑 Browser closed. Saving Excel...")

        df.to_excel(output_excel_path, index=False)

        try:
            wb = load_workbook(output_excel_path)
            ws = wb.active

            # Header Styling
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                for cell in col_cells:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    width = max(len(str(cell.value)) + 5, 15)
                    ws.column_dimensions[get_column_letter(col_num)].width = width

            # Data Cell Styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(output_excel_path)
            print(f"✅ Excel styled and saved at: {output_excel_path}")

        except Exception as e:
            print(f"⚠️ Error styling Excel: {e}")

# ----------------------------
# ✅ Run the Function
# ----------------------------
if __name__ == "__main__":
    input_excel_path = "Pubs.xlsx"           # Ensure this file exists
    output_excel_path = "Pubs_Updated.xlsx"
    process_excel(input_excel_path, output_excel_path)
